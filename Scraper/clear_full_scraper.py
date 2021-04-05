import os
from bs4 import BeautifulSoup
import re
from openpyxl.workbook import Workbook


directory = "C:\\Users\\P\\PycharmProjects\\my_own_projects_and_ideas\\Обзор акций html"   #УКАЗЫВАЕМ ПУТЬ К ПАПКЕ, ГДЕ ЛЕЖАТ ВСЕ СКАЧАННЫЕ СТАРНИЦЫ
files = os.listdir(directory)      #ТУТ ПРОХОДИМ ПО УКАЗАННОЙ ПАПКЕ И СОБИРАЕМ ВСЕ НАЗВАНИЯ ФАЙЛОВ В СПИСОК files
s = filter(lambda x: x.endswith('.html'), files)     #ОСТАВЛЯЕМ ТОЛКЬКО ТЕ, У КОТОРЫХ РАСШИРЕНИЕ .html
html = [*s]         #И СОБИРАЕМ ИХ В СПИСОК

headers = ['title', 'cat', 'Market Cap\nСтоимость компании', 'Ebitda\nПрибыль до вычета расходов',
           'P/E\nЦена акции / прибыль', 'P/S\nЦена акции / выручка', 'Diluted EPS\nРазводненная прибыль на акцию',
           'Рост EPS\nСредний рост за 5 лет', 'Рост выручки\nСредний рост за 5 лет', 'ROE\nДоходность капитала',
           'ROA\nДоходность активов', 'ROI\nДоходность инвестиций', 'Debt/Equity\nДолг / Капитал',
           'Net Profit Margin\nПрибыль в % от выручки', 'Payout Ratio\nПроцент дивидендов от прибыли',
           'Средний дивидендный доход\nЗа 5 лет', 'Дивидендная доходность\nЗа год', 'Цена открытия\n\xa0',
           'Цена закрытия\n\xa0', '52 w High\nМаксимум за год', '52 w Low\nМинимум за год',
           'Дневной объем торгов\nСредний за 10 дней', 'Месячный объем торгов\nСредний за 3 месяца',
           'Beta\nВолатильности акций к рынку']       #ЭТО СПИСОК С НАЗВАНИЯМИ КОЛОНОК (Я ЕГО СОБРАЛ ЗАРАНЕЕ)


#СОЗДАНИЕ ШАПКИ И EXCEL ФАЙЛА
name = 'name.xlsx'
workbook_name = name
wb = Workbook()
page = wb.active
page.title = 'investments_companies'
page.append(headers)            # write the headers to the first line
wb.save(filename = workbook_name)
# БУДЬТЕ ВНИМАТЕЛЬНЫ, ЭТОТ БЛОК (СТРОКИ 24-30) СОЗДАЕТ НОВЫЙ ФАЙЛ,
# ЕСЛИ ХОТИТЕ ДОБАВИТЬ К УЖЕ СУЩЕСТВУЮЩЕМУ ФАЙЛУ ДАННЫЕ, ЗАКОММЕНТИРУЙТЕ ЭТОТ БЛОК,
# А В СТРОКЕ 66 УКАЖИТЕ ИМЯ ФАЙЛА В КОТОРЫЙ ХОТИТЕ ДОПИСАТЬ


full_path = 'C:\\Users\\P\\PycharmProjects\\my_own_projects_and_ideas\\Обзор акций html\\'  #ПУТЬ СО "\\"" В КОНЦЕ
#НЕПОСРЕДСТВЕННО САМ СКРЭППИНГ
for adr in html:                                                    #ПЕРЕБИРАЕМ ВСЕ НАЗВАНИЯ ФАЙЛОВ ИЗ СПИСКА html
    with open (full_path+adr, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f, 'html.parser')                      #ОТКРЫВАЕМ И ПАРСИМ
    data = soup.find_all('td', class_='Table__cell_3dA6T')          # ПОИСК ВСЕХ ЧИСЛЕННЫХ ДАННЫХ НА СТРАНИЦЕ
    category = soup.find_all('div', class_="SecurityHeaderPure__panelText_2m2wp")    # ПОСИК "СЕКТОРА ЭКОНОМИКИ"
    cat = [n.get_text() for n in category]
    t = re.findall('(Акции )(.*)( \()', soup.find('title').get_text())               # НАЗВАНИЕ АКЦИИ
    li = ['title', t[0][1], 'cat', cat[1]]
    for n in data:
        li.append(n.get_text('\n'))
    d = {}
    k = 0
    v = 1
    while v <= len(li):
        d[li[k]] = li[v]
        k += 2
        v += 2
    #print(d)     #ЭТО ЕСЛИ ЗАХОЧЕТСЯ ПОСМОТРЕТЬ ВЫВОД ДАННЫХ

    # ФОРМИРОВАНИЕ СПИСКА С ВЫХОДНЫМИ ДАННЫМИ
    values = []
    for k in headers:
        c = d.get(k, '0')    #ЕСЛИ ЗНАЧЕНИЯ В СЛОВАРЕ "d" НЕ НАЙДЕНО, ПРИСВОИТЬ "0"
        a = re.sub('\n', '', c)     #УБИРАЕМ \n ПОСЛЕ get_text('\n')
        values.append(a)

    # ДОБАВЛЕНИЕ НОВЫХ ДАННЫХ
    from openpyxl import load_workbook
    workbook_name = name
    wb = load_workbook(workbook_name)
    page = wb.active
    page.append(values)
    values.clear()
    wb.save(filename=workbook_name)